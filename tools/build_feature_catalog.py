#!/usr/bin/env python3
"""
Build the JustType Feature Catalog spreadsheet for collaborator review.

Inputs:
  SETTINGS_INFO_PROMPTS.txt  — single source of truth for pages/sections/settings/INFO PROMPTs.
  docs/Phase1SettingsWork.md — surface coverage, discrepancy flags, renames (best-effort regex extraction).

Output:
  docs/JustType_Feature_Catalog.xlsx — six-tab workbook ready for Drive upload.

Run from repo root:
  tools/.venv/bin/python tools/build_feature_catalog.py
"""

from __future__ import annotations

import re
import sys
from collections import OrderedDict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCE_TXT = REPO_ROOT / "SETTINGS_INFO_PROMPTS.txt"
PHASE1_MD = REPO_ROOT / "docs" / "Phase1SettingsWork.md"
# Bump this constant when the catalog schema changes materially (added/removed
# columns, new sheets, etc.) so reviewers can tell which revision they're
# looking at. v3 (2026-05-04): added the "Disclosure" column for user-journey
# tier (Essential / Standard / Intermediate / Advanced).
OUTPUT_XLSX = REPO_ROOT / "docs" / "JustType_Feature_Catalog.v3.xlsx"


# ── Page-prefix mapping (used to generate code names) ──────────────────────
# Order here defines presentation order in the Code Name Legend tab.
PAGE_PREFIXES: "OrderedDict[str, str]" = OrderedDict([
    ("JustType Settings (Main)", "M"),
    ("Input Methods", "IM"),
    ("Set Up Head Tracking", "HT"),
    ("Set up Joystick", "JS"),
    ("Set up Single Switch Scanning", "SS"),
    ("Two Switch Selection Set-Up", "TS"),
    ("Set up Touchscreen Switch", "TC"),
    ("Direct Selection", "DS"),
    ("Directional Selection", "DR"),
    ("Vocabulary Management", "VM"),
    ("Select Active Vocabularies", "SV"),
    ("Export Vocabulary Usage", "EV"),
    ("Manage Vocabularies", "MV"),
    ("Backup & Restore", "BR"),
    ("(Infrastructure)", "INF"),
])


# ── Tier defaults (proposed, Kirk overrides) ───────────────────────────────
# T1 = Hard-gate (potentially destructive or "scary" actions)
# T2 = Visible-but-locked (complex setup that benefits from gating)
# T3 = Open (safe, day-to-day adjustments)
T1_KEYWORDS = [
    "restore from backup",
    "delete selected vocabularies",
    "merge selected vocabularies",
    "use advanced backup",
]
T2_KEYWORDS = [
    "active zone",
    "exit zone",
    "resting zone",
    "key activation threshold",
    "corner bias",
    "corner weighting",
    "vertical sensitivity",
    "responsiveness",
    "response curve",
    "debug overlay",
    "stuck switch",
    "switch debounce",
    "touch debounce",
    "swipe debounce",
    "auto-repeat",
    "scan step delay",
    "scan repeat count",
    "highlight timeout",
    "two-switch selection reset delay",
    "exit/pause keyboard delay",
    "overlay timeout",
    "overlay mode display timeout",
    "switch mode",
    "touchscreen switch height",
    "touchscreen switch mode",
]


PAGE_DEFAULT_TIER = {
    "Set Up Head Tracking": "T2",
    "Set up Joystick": "T2",
    "Set up Single Switch Scanning": "T2",
    "Two Switch Selection Set-Up": "T2",
    "Set up Touchscreen Switch": "T2",
    "Manage Vocabularies": "T2",
}


def propose_tier(setting_name: str, page: str) -> str:
    name = setting_name.lower()
    # Strongest signals: explicit T1 keywords (destructive operations).
    for kw in T1_KEYWORDS:
        if kw in name:
            return "T1"
    # Specific T2 keywords override page defaults upward (no effect when page is already T2).
    for kw in T2_KEYWORDS:
        if kw in name:
            return "T2"
    if "assign" in name and "switch" in name:
        return "T2"
    # Otherwise inherit page default; fall back to T3.
    return PAGE_DEFAULT_TIER.get(page, "T3")


# ── Disclosure tier (v3, 2026-05-04) ──────────────────────────────────────
# A 4-level user-journey progression that asks: "at what point in the user's
# journey should this control become visible?" Distinct from Proposed Tier
# (which is about gating policy on the keyboard side). Every row gets a
# single-letter recommendation; Kirk overrides via the narrow Disclosure
# column on the Features sheet.
#
#   E = Essential    — visible to first-time users; cannot be hidden.
#                      The fundamental decisions a user must make to type at all
#                      (input method choice, key feedback, basic speech).
#   S = Standard     — visible by default to everyone; might be optionally
#                      hidden later if we add a Simplify Mode. Common day-to-day
#                      adjustments (keyboard size, predictions, basic backup).
#   I = Intermediate — hidden until the user opts into "Show More Options".
#                      Power-user adjustments (debounce, auto-repeat, vocab
#                      promotion, capitalization timing).
#   A = Advanced     — hidden behind an explicit "Show Advanced Settings"
#                      master toggle (List B). Calibration sliders, frequency-
#                      class thresholds, debug overlays, advanced backup.

# Exact-match overrides (full feature_name == this string, case-insensitive).
# Used where substring matches would over-trigger (e.g. "touchscreen switch"
# as an Adaptive method name vs. "Use Touchscreen Switch" elsewhere).
DISCLOSURE_ESSENTIAL_EXACT = {
    "letter arrangement",
    "head tracking",
    "joystick",
    "single switch scanning",
    "two-switch selection",
    "direct selection",
    "directional selection",
    "touchscreen switch",
}

DISCLOSURE_ESSENTIAL_SUBSTRING = [
    "beep when key is activated",
    "beep when keystroke has no effect",
    "flash key when activated",
    "speak each output word",
    "create backup now",
    "info-icon mechanism",
    "show setting help",
    "touchscreen input methods",
    "adaptive input methods",
    "auditory feedback",
    "visual feedback",
    "how touchscreen switch works",
]

DISCLOSURE_ADV_KEYWORDS = [
    "active zone",
    "resting zone",
    "exit zone",
    "key activation threshold",
    "vertical sensitivity",
    "responsiveness",
    "response curve",
    "corner weighting",
    "corner bias",
    "exit/pause keyboard delay",
    "minimum frequency",
    "maximum frequency",
    "min freq",
    "max freq",
    "frequency for words",
    "frequency to show accented",
    "stuck switch",
    "overlay mode",  # both "Overlay Mode" and "Overlay Mode Display Timeout"
    "debug overlay",
    "use advanced",
    "merge selected vocab",
    "delete selected vocab",
    "developer",
    "include advanced",
    "backup history",
    "pre-populated backup",
]

DISCLOSURE_INT_KEYWORDS = [
    "debounce",
    "auto-repeat",
    "auto repeat",
    "wait until word",
    "delay to add capitalized",
    "case type",
    "deferred",
    "case expansion",
    "highlight timeout",
    "two-switch selection reset delay",
    "promote",
    "restricted words",
    "restrict uncommon",
    "auto-speak delay",
    "phrase speak delay",
    "speak the name of punctuation",
    "speak punctuation",
    "shrink",
    "key history height",
    "use touchscreen switch",
    "import vocabulary",
    "export usage",
    "include phrases",
    "skip keys",
    "show next key",
    "select key triggers",
    "repeat scan",
    "extra delay on first",
    "first scanned",
    "switch mode",
    "swipe debounce",
    "swipe sensitivity",
    "manage justtype database",
    "abbreviation",
    "phrase in selection",
    "key history",
    "color code band",
    "scan layout",
    "scan step delay",
    "overlay opacity",
    "scanning layout key size",
]

DISCLOSURE_PAGE_DEFAULT = {
    "JustType Settings (Main)": "S",
    "Input Methods": "S",
    "Set Up Head Tracking": "A",
    "Set up Joystick": "A",
    "Set up Single Switch Scanning": "S",
    "Two Switch Selection Set-Up": "S",
    "Set up Touchscreen Switch": "S",
    "Direct Selection": "S",
    "Directional Selection": "S",
    "Vocabulary Management": "S",
    "Select Active Vocabularies": "S",
    "Export Vocabulary Usage": "I",
    "Manage Vocabularies": "I",
    "Backup & Restore": "S",
    "(Infrastructure)": "I",
}


def propose_disclosure(name: str, page: str) -> str:
    """Recommend an Essential / Standard / Intermediate / Advanced disclosure
    level for a feature. Kirk overrides per row on the Features sheet.

    Order of precedence:
      1. Exact-name match against ESSENTIAL list (most specific).
      2. Substring match against ESSENTIAL list.
      3. Substring match against ADV (advanced) keywords.
      4. Substring match against INT (intermediate) keywords.
      5. Page default (most pages default to S; calibration pages default to A).
    """
    blob = name.lower().strip()
    # Strip our own annotation prefixes ("[Section]", "[Page-level]", etc.).
    blob = re.sub(r"^\[(?:section|page-level|column|row)\]\s*", "", blob)

    if blob in DISCLOSURE_ESSENTIAL_EXACT:
        return "E"
    for kw in DISCLOSURE_ESSENTIAL_SUBSTRING:
        if kw in blob:
            return "E"
    for kw in DISCLOSURE_ADV_KEYWORDS:
        if kw in blob:
            return "A"
    for kw in DISCLOSURE_INT_KEYWORDS:
        if kw in blob:
            return "I"
    return DISCLOSURE_PAGE_DEFAULT.get(page, "S")


# ── Audience guesses (proposed, blank where unsure) ────────────────────────
# Triggered by patterns where strategic guidance is plausibly role-specific.
TEACHER_KEYWORDS = [
    "accent",
    "frequency",
    "vocabulary module",
    "import",
    "merge",
    "promote imported",
]
CLINICIAN_KEYWORDS = [
    "head tracking",
    "joystick",
    "single-switch",
    "two-switch",
    "switch debounce",
    "active zone",
    "exit zone",
    "corner bias",
    "corner weighting",
    "vertical sensitivity",
    "responsiveness",
]


def propose_audience(page: str, setting: str) -> str:
    audiences: list[str] = ["User"]
    blob = f"{page} {setting}".lower()
    if any(k in blob for k in TEACHER_KEYWORDS):
        audiences.append("Teacher")
    if any(k in blob for k in CLINICIAN_KEYWORDS):
        audiences.append("Clinician")
    if len(audiences) == 1:
        # Defaults are always "User" alone — leave blank so Kirk's eye is drawn
        # to rows where we did flag a non-default audience.
        return ""
    return ", ".join(audiences)


# ── Hand-curated overrides for known discrepancies / new features ──────────
# Maps either a NEW SETTING NAME (preferred) or current SETTING NAME → status.
#
# Updated 2026-05-01 to reflect Phase 3B + 3C + 3D-Tier-1 implementation
# progress. Items removed from the sets default to "Implemented per INFO
# PROMPT" via the parser (or "Renamed only" if the original differs).
KNOWN_BEHAVIOR_GAPS = {
    # Auto-Speak Delay (formerly Phrase Speak Delay): the Phase 1 ledger
    # flagged this as "current code treats slider max as OFF; spec requires
    # 0 = OFF". Slider position 0 = 0 ms = no auto-speak is already the
    # current behavior; the broader 0=OFF/1..20 reframing has not been
    # explicitly verified against the ms-based slider yet, so still flagged.
    "auto-speak delay",
    "phrase speak delay",
    # Δ-5 Direct/Directional minimum-one constraint relaxation — Phase 3D
    # Tier 2 (still open).
    "direct selection",
    "directional selection",
}

KNOWN_NOT_YET_IMPLEMENTED = {
    # Phase 3C-C / 3C-D pending: brand-new settings not yet wired.
    "optimized letter arrangement language",
    "exit zone (active zone < exit zone < 1.0)",
    "exit/pause keyboard delay time (2 - 5 sec)",
    # Backup & Restore advanced flow — Phase 3E.
    "use advanced backup & restore",
    "[page-level] previous backups",
    "[page-level] choose/create new backup folder",
    # Touchscreen Switch Height: the System UI slider exists but the IME's
    # rendering pipeline still uses the old fixed height. Flag as
    # not-yet-implemented end-to-end. (Earlier ledger-only entry —
    # confirm during Phase 3D Tier 2/3.)
    # "touchscreen switch height",  -- UI exists, behavior verification pending
    # Removed (now implemented):
    #   "active zone (...). maximum value for active zone is 0.75"
    #     -- HT slider already capped at 0.75; only label was renamed.
    #   "touchscreen switch height" -- UI/pref both wired (Phase 3C-A).
    #   "overlay mode" -- UI/pref wired; rendering still pending but
    #                     the catalog row tracks the setting, not the
    #                     underlying renderer; flag the renderer
    #                     separately if needed.
}

# Phase 3D Tier 1 closures (Δ-6, Δ-13, Δ-35) and any other resolved
# discrepancies — removed from the gaps set above so the catalog status
# falls through to "Implemented per INFO PROMPT" or "Renamed only".


# ── Hand-curated dependencies (proposed, blank elsewhere) ──────────────────
DEPENDENCIES = {
    "head tracking": "HeadBoard companion app installed",
    "joystick": "External joystick + HeadBoard switch",
    "single-switch": "Bluetooth or HeadBoard switch",
    "two-switch": "Two Bluetooth or HeadBoard switches",
    "use touchscreen switch": "Touchscreen Switch enabled on Input Methods page",
    "promote imported": "At least one Imported Vocabulary loaded",
    "include phrases": "Abbreviation/Phrase pairs added via Add New Phrase",
    "use advanced backup": "Backup history infrastructure (Phase 3E)",
    "restore from backup": "At least one completed backup exists",
}


def propose_dependency(setting: str, page: str) -> str:
    blob = f"{page} {setting}".lower()
    for kw, dep in DEPENDENCIES.items():
        if kw in blob:
            return dep
    return ""


# ── Hand-curated infrastructure rows (List B + Phase 3E additions) ─────────
INFRASTRUCTURE_ROWS = [
    {
        "feature_name": "Include Advanced Settings in Keyboard Settings (master toggle)",
        "page": "(Infrastructure)",
        "section": "Sensitive-Settings Gating",
        "info_prompt": (
            "Master toggle (default OFF) that exposes Advanced settings inside Keyboard "
            "Settings. When OFF, advanced/sensitive items are hidden from the keyboard-side "
            "navigation; when ON, they appear in their normal positions but with confirm-by-"
            "test-mode semantics where appropriate. Safe day-to-day controls (T3) are unaffected."
        ),
        "tier": "T2",
        "advanced_affiliation": "(parent)",
        "status": "Not yet implemented",
        "tutorial_topic": "Setting up JustType for someone else",
        "audience": "Teacher, Clinician",
        "dependencies": "",
    },
    {
        "feature_name": "Backup History (multi-folder backup manifest)",
        "page": "(Infrastructure)",
        "section": "Backup History",
        "info_prompt": (
            "Replaces the single backup_tree_uri preference with a list of every backup folder "
            "JustType has used, each with a last-backup timestamp. Powers the 'Previous Backups' "
            "list under Use Advanced Backup & Restore."
        ),
        "tier": "T2",
        "advanced_affiliation": "Use Advanced Backup & Restore",
        "status": "Not yet implemented",
        "tutorial_topic": "Backup & Restore for advanced users",
        "audience": "Teacher, Clinician",
        "dependencies": "BackupManager API extension",
    },
    {
        "feature_name": "Pre-populated Backup Slots (5 default slots on install)",
        "page": "(Infrastructure)",
        "section": "Backup History",
        "info_prompt": (
            "5 named backup slots are pre-populated at install time so Restore can be performed "
            "from the keyboard without invoking SAF. User assigns folders to slots from System "
            "Settings; keyboard-side picks among already-assigned slots."
        ),
        "tier": "T1",
        "advanced_affiliation": "Use Advanced Backup & Restore",
        "status": "Not yet implemented",
        "tutorial_topic": "Backup & Restore for advanced users",
        "audience": "Teacher, Clinician",
        "dependencies": "Backup History",
    },
    {
        "feature_name": "Pre-installed Vocabulary Module Catalog",
        "page": "(Infrastructure)",
        "section": "Vocabulary Catalog",
        "info_prompt": (
            "Curated set of pre-installed Vocabulary Modules (Medical, Legal, Programming, "
            "Cooking, Academic-Biochem, etc.) shipped with the app so advanced users can enable "
            "specialty vocabularies from the keyboard without importing files."
        ),
        "tier": "T3",
        "advanced_affiliation": "",
        "status": "Not yet implemented",
        "tutorial_topic": "Vocabulary Modules for advanced users",
        "audience": "User, Teacher",
        "dependencies": "",
    },
    {
        "feature_name": "Info-icon mechanism (System Settings)",
        "page": "(Infrastructure)",
        "section": "INFO PROMPT exposure",
        "info_prompt": (
            "Reusable helper (SettingsInfoHelper) that attaches a small 'i' icon next to any "
            "System Settings control. Tapping the icon shows the INFO PROMPT in a dialog and, "
            "when 'Speak Prompts Aloud' is ON, speaks the prompt via SettingsSpeechController. "
            "Speech cancels on dialog dismiss and on activity onPause. Footnote-style icon "
            "placement (translationY = -8dp) with a TouchDelegate-expanded hit region. "
            "Implemented in Phase 3B and propagated across all ~14 settings activities."
        ),
        "tier": "T3",
        "advanced_affiliation": "",
        "status": "Implemented per INFO PROMPT",
        "tutorial_topic": "Discovering settings help",
        "audience": "User, Teacher, Clinician",
        "dependencies": "",
    },
    {
        "feature_name": "SHOW SETTING HELP / SHOW SECTION HELP keys (Keyboard Settings)",
        "page": "(Infrastructure)",
        "section": "INFO PROMPT exposure",
        "info_prompt": (
            "Keys 3 and 4 in PAGE_NAV mode become SHOW SETTING HELP / SHOW SECTION HELP when an "
            "INFO PROMPT exists for the focused setting/section. Section help cycles ('SHOW SECTION "
            "HELP 1', '2', ...) when multiple prompts apply. Activation pops a dismiss-on-next-key "
            "panel and speaks the prompt when Speak Prompts Aloud is ON."
        ),
        "tier": "T3",
        "advanced_affiliation": "",
        "status": "Not yet implemented",
        "tutorial_topic": "Discovering settings help",
        "audience": "User, Teacher, Clinician",
        "dependencies": "SettingsDef.infoPrompt field",
    },
]


# ── Source-file parser ─────────────────────────────────────────────────────

PAGE_LINE = re.compile(r"^\s*PAGE(?:\s+NAME)?\s*[:=]\s*(.+?)\s*$", re.IGNORECASE)
NEW_PAGE_LINE = re.compile(r"^\s*NEW\s+PAGE\s+NAME\s*[:=]\s*(.*?)\s*$", re.IGNORECASE)
SECTION_LINE = re.compile(r"^\s*SECTION\s+NAME\s*[:=]\s*(.+?)\s*$", re.IGNORECASE)
NEW_SECTION_LINE = re.compile(r"^\s*NEW\s+SECTION\s+NAME\s*[:=]\s*(.*?)\s*$", re.IGNORECASE)
SETTING_LINE = re.compile(r"^\s*SETTING\s*(?:\(.*?\))?\s*NAME\s*[:=]\s*(.+?)\s*$", re.IGNORECASE)
NEW_SETTING_LINE = re.compile(r"^\s*NEW\s+SETTING\s*(?:\(.*?\))?\s*NAME\s*[:=]\s*(.*?)\s*$", re.IGNORECASE)
COLUMN_HEADING_LINE = re.compile(r"^\s*COLUMN\s+HEADING\s*[:=]\s*(.+?)\s*$", re.IGNORECASE)
NEW_COLUMN_HEADING_LINE = re.compile(r"^\s*NEW\s+COLUMN\s+HEADING\s*[:=]\s*(.*?)\s*$", re.IGNORECASE)
ROW_HEADING_LINE = re.compile(r"^\s*ROW\s+HEADING\s*[:=]\s*(.+?)\s*$", re.IGNORECASE)
NEW_ROW_HEADING_LINE = re.compile(r"^\s*NEW\s+ROW\s+HEADING\s*[:=]\s*(.*?)\s*$", re.IGNORECASE)
INFO_PROMPT_LINE = re.compile(r"^\s*INFO\s+PROMPT[^:]*[:=]\s*(.*)$", re.IGNORECASE)
INFO_PROMPT_PARENTHETICAL = re.compile(
    r"^\s*INFO\s+PROMPT\s*\((.*?)\)\s*[:=]\s*(.*)$", re.IGNORECASE
)


@dataclass
class FeatureRow:
    code_name: str = ""
    feature_name: str = ""
    page: str = ""
    section: str = ""
    surface: str = "Both"
    status: str = "Implemented per INFO PROMPT"
    tier: str = ""
    disclosure: str = ""  # E / S / I / A — see propose_disclosure().
    advanced_affiliation: str = ""
    settings_group: str = ""
    tutorial_topic: str = ""
    audience: str = ""
    dependencies: str = ""
    info_prompt: str = ""

    # Internal: original (un-renamed) name for legacy lookups.
    original_name: str = ""


def clean(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def strip_section_dashes(name: str) -> str:
    """Remove leading/trailing '---' decorations and whitespace."""
    return name.strip().strip("-").strip()


def parse_source(text: str) -> list[FeatureRow]:
    rows: list[FeatureRow] = []
    lines = text.splitlines()

    current_page = ""
    current_section = ""
    pending_setting: Optional[FeatureRow] = None
    info_buf: list[str] = []
    in_info = False

    def flush() -> None:
        nonlocal pending_setting, info_buf, in_info
        if pending_setting is None:
            return
        pending_setting.info_prompt = clean(" ".join(info_buf))
        rows.append(pending_setting)
        pending_setting = None
        info_buf = []
        in_info = False

    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        # ── Page boundary (lines of '=') ─────────────────────────────
        if set(stripped) == {"="} and len(stripped) >= 20:
            # Look for next non-blank line that gives a page name.
            j = i + 1
            while j < len(lines) and not lines[j].strip():
                j += 1
            if j < len(lines):
                m = PAGE_LINE.match(lines[j])
                if m:
                    flush()
                    current_page = clean(m.group(1))
                    # Look for NEW PAGE NAME on subsequent line.
                    if j + 1 < len(lines):
                        nm = NEW_PAGE_LINE.match(lines[j + 1])
                        if nm and nm.group(1).strip():
                            current_page = clean(nm.group(1))
                            i = j + 2
                        else:
                            i = j + 1
                    else:
                        i = j + 1
                    current_section = ""
                    continue
            i += 1
            continue

        # ── Section ─────────────────────────────────────────────────
        msec = SECTION_LINE.match(line)
        if msec:
            flush()
            sec = clean(msec.group(1))
            # Try NEW SECTION on the next line.
            if i + 1 < len(lines):
                nm = NEW_SECTION_LINE.match(lines[i + 1])
                if nm and nm.group(1).strip():
                    sec = clean(nm.group(1))
                    i += 2
                else:
                    i += 1
            else:
                i += 1
            current_section = strip_section_dashes(sec)
            continue

        # ── Setting / Column Heading / Row Heading ──────────────────
        mset = SETTING_LINE.match(line)
        mcol = COLUMN_HEADING_LINE.match(line)
        mrow = ROW_HEADING_LINE.match(line)
        if mset or mcol or mrow:
            flush()
            if mset:
                current_name = clean(mset.group(1))
                new_pat, kind = NEW_SETTING_LINE, "setting"
            elif mcol:
                current_name = clean(mcol.group(1))
                new_pat, kind = NEW_COLUMN_HEADING_LINE, "column"
            else:
                current_name = clean(mrow.group(1))
                new_pat, kind = NEW_ROW_HEADING_LINE, "row"

            new_name = ""
            if i + 1 < len(lines):
                nm = new_pat.match(lines[i + 1])
                if nm:
                    new_name = clean(nm.group(1))
                    i += 2
                else:
                    i += 1
            else:
                i += 1

            # Skip placeholder/non-real entries.
            placeholder = (
                current_name.startswith("[") and current_name.rstrip().endswith("]")
            )
            display_name = new_name if new_name else current_name
            if placeholder and not new_name:
                continue

            # Annotate non-setting kinds so Kirk can spot column/row entries.
            if kind == "column":
                display_name = f"[Column] {display_name}"
            elif kind == "row":
                display_name = f"[Row] {display_name}"

            row = FeatureRow(
                feature_name=display_name,
                page=current_page,
                section=current_section,
                original_name=current_name if not placeholder else "",
            )
            pending_setting = row
            info_buf = []
            in_info = False
            continue

        # ── Parenthetical INFO PROMPT (page- or table-level) ────────
        # E.g. `INFO PROMPT (for "Active Word Count"): ...`
        mpar = INFO_PROMPT_PARENTHETICAL.match(line)
        if mpar and pending_setting is None:
            flush()
            label = clean(mpar.group(1))
            # Strip noise words to leave a usable display name.
            cleaned_label = re.sub(r'^(for|next to|shown above|shown beneath|shown at the top of|at the top of|at the top of the)\s+', '', label, flags=re.IGNORECASE)
            cleaned_label = cleaned_label.strip().strip('"').strip()
            # Quote-bracketed names inside the parenthetical are the most informative.
            qmatch = re.search(r'"([^"]+)"', label)
            if qmatch:
                cleaned_label = qmatch.group(1)
            if not cleaned_label:
                cleaned_label = label
            row = FeatureRow(
                feature_name=f"[Page-level] {cleaned_label}",
                page=current_page,
                section=current_section,
                original_name="",
            )
            pending_setting = row
            info_buf = []
            tail = mpar.group(2).strip()
            if tail:
                info_buf.append(tail)
            in_info = True
            i += 1
            continue

        # ── INFO PROMPT (start of multi-line block) ─────────────────
        minfo = INFO_PROMPT_LINE.match(line)
        if minfo:
            in_info = True
            tail = minfo.group(1).strip()
            if tail:
                info_buf.append(tail)
            i += 1
            continue

        # ── Continue accumulating INFO PROMPT body ──────────────────
        if in_info:
            if not stripped:
                # Blank line ends the prompt.
                if pending_setting is not None:
                    flush()
                else:
                    in_info = False
                    info_buf = []
            else:
                # Skip meta-instruction blocks inside an INFO PROMPT.
                if stripped.startswith(">>>") or stripped.startswith("<<<"):
                    i += 1
                    continue
                # End the prompt if a new structural marker appears.
                if (
                    SETTING_LINE.match(line)
                    or SECTION_LINE.match(line)
                    or PAGE_LINE.match(line)
                    or set(stripped) == {"="}
                    or set(stripped) == {"-"}
                ):
                    flush()
                    continue  # re-process this line in the outer loop
                info_buf.append(stripped)
            i += 1
            continue

        i += 1

    flush()
    return rows


# ── Phase 1 surface/status overlay (best-effort regex) ─────────────────────

def overlay_phase1_data(rows: list[FeatureRow]) -> None:
    """Best-effort: scan Phase1SettingsWork.md for 'Δ-N' rows tied to setting names.

    Keeps the overlay conservative — only marks "Behavior gap" when the row's
    feature name appears near a Δ-marker. Surface coverage stays at the default
    'Both' unless a stronger signal is found.
    """
    if not PHASE1_MD.exists():
        return
    text = PHASE1_MD.read_text(encoding="utf-8", errors="replace").lower()

    for row in rows:
        if row.page == "(Infrastructure)":
            continue
        key = row.feature_name.lower()
        if not key or len(key) < 5:
            continue
        # Crude proximity: setting name occurs within ~120 chars of a Δ- token.
        idx = 0
        while True:
            j = text.find(key, idx)
            if j == -1:
                break
            window = text[max(0, j - 200): j + 200]
            if "Δ-" in window or "delta-" in window or "behavior gap" in window:
                row.status = "Behavior gap — needs code update"
                break
            idx = j + len(key)


def apply_known_overrides(rows: list[FeatureRow]) -> None:
    for row in rows:
        key = row.feature_name.lower().strip()
        orig = row.original_name.lower().strip()
        if key in KNOWN_BEHAVIOR_GAPS or orig in KNOWN_BEHAVIOR_GAPS:
            row.status = "Behavior gap — needs code update"
        if key in KNOWN_NOT_YET_IMPLEMENTED or orig in KNOWN_NOT_YET_IMPLEMENTED:
            row.status = "Not yet implemented"
        # Renamed-only signal: no status overlay above; original differs and INFO
        # PROMPT only adjusts wording.
        if (
            row.original_name
            and row.original_name.strip().lower() != row.feature_name.strip().lower()
            and row.status == "Implemented per INFO PROMPT"
        ):
            row.status = "Renamed only"


# ── Code-name assignment ───────────────────────────────────────────────────

def assign_code_names(rows: list[FeatureRow]) -> None:
    counters: dict[str, int] = {}
    for row in rows:
        prefix = PAGE_PREFIXES.get(row.page, "X")
        counters[prefix] = counters.get(prefix, 0) + 1
        row.code_name = f"{prefix}-{counters[prefix]}"


# ── Defaults for proposal columns ──────────────────────────────────────────

def fill_proposals(rows: list[FeatureRow]) -> None:
    for row in rows:
        # Disclosure runs for every row (including infrastructure) — those
        # rows have explicit overrides on a few specific items, fallthrough
        # for the rest.
        if not row.disclosure:
            row.disclosure = propose_disclosure(row.feature_name, row.page)
        if row.page == "(Infrastructure)":
            # Other proposal columns are already set on infrastructure rows.
            row.settings_group = row.settings_group or row.section or row.page
            continue
        row.tier = row.tier or propose_tier(row.feature_name, row.page)
        row.audience = row.audience or propose_audience(row.page, row.feature_name)
        row.dependencies = row.dependencies or propose_dependency(row.feature_name, row.page)
        row.settings_group = row.settings_group or (row.section if row.section else row.page)
        row.tutorial_topic = row.tutorial_topic or row.page


# ── Workbook construction ──────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)


def autosize(ws, max_widths: dict[int, int]) -> None:
    for col_idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_workbook(rows: list[FeatureRow]) -> Workbook:
    wb = Workbook()

    # ── Sheet 1: Features ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Features"
    headers = [
        "Code Name", "Feature Name", "Page", "Section", "Surface",
        "Status", "Proposed Tier",
        # v3 (2026-05-04): Disclosure tier (E/S/I/A). Single-letter values
        # so the column stays narrow; full definitions on the Glossary tab.
        "Disclosure",
        "Advanced Flag-Affiliation",
        "Proposed Settings Group", "Proposed Tutorial Topic",
        "Audience", "Dependencies",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Sort: Page (in PAGE_PREFIXES order), then Section (document order via index map).
    page_order = {p: i for i, p in enumerate(PAGE_PREFIXES.keys())}
    section_order: dict[tuple[str, str], int] = {}
    seen_idx = 0
    for r in rows:
        key = (r.page, r.section)
        if key not in section_order:
            section_order[key] = seen_idx
            seen_idx += 1
    rows_sorted = sorted(
        rows,
        key=lambda r: (
            page_order.get(r.page, len(page_order)),
            section_order.get((r.page, r.section), 9999),
        ),
    )

    # Map code_name → INFO PROMPTS row index for HYPERLINK.
    info_row_index: dict[str, int] = {}
    for idx, r in enumerate(rows_sorted, start=2):  # row 2 onward in INFO PROMPTS
        info_row_index[r.code_name] = idx

    for r in rows_sorted:
        ws.append([
            r.code_name,
            r.feature_name,
            r.page,
            r.section,
            r.surface,
            r.status,
            r.tier,
            r.disclosure,
            r.advanced_affiliation,
            r.settings_group,
            r.tutorial_topic,
            r.audience,
            r.dependencies,
        ])
    # Attach internal hyperlinks on Code Name cells (column A).
    for excel_row, r in enumerate(rows_sorted, start=2):
        cell = ws.cell(row=excel_row, column=1)
        cell.hyperlink = f"#'INFO PROMPTS'!A{info_row_index[r.code_name]}"
        cell.font = Font(color="0563C1", underline="single")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    # Column widths — column 8 (Disclosure) is intentionally narrow (~6 chars)
    # so the new tier sits as a slim band next to the existing Tier column.
    autosize(
        ws,
        {1: 12, 2: 38, 3: 24, 4: 28, 5: 12, 6: 26, 7: 8, 8: 6,
         9: 22, 10: 26, 11: 26, 12: 18, 13: 26},
    )
    for row_cells in ws.iter_rows(min_row=2):
        for c in row_cells:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    # Center the narrow Disclosure column for readability.
    for excel_row in range(2, ws.max_row + 1):
        ws.cell(row=excel_row, column=8).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=False,
        )

    # Data validations on Surface/Status/Tier/Disclosure/Audience.
    dv_surface = DataValidation(type="list", formula1='"System,Keyboard,Both,Neither"', allow_blank=True)
    dv_status = DataValidation(
        type="list",
        formula1='"Implemented per INFO PROMPT,Behavior gap — needs code update,Not yet implemented,Renamed only"',
        allow_blank=True,
    )
    dv_tier = DataValidation(type="list", formula1='"T1,T2,T3"', allow_blank=True)
    dv_disclosure = DataValidation(type="list", formula1='"E,S,I,A"', allow_blank=True)
    last_row = ws.max_row
    for dv, col in [(dv_surface, 5), (dv_status, 6), (dv_tier, 7), (dv_disclosure, 8)]:
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(col)}2:{get_column_letter(col)}{last_row}")

    # ── Sheet 2: INFO PROMPTS ────────────────────────────────────────
    ws2 = wb.create_sheet("INFO PROMPTS")
    ws2.append(["Code Name", "Feature Name", "Page › Section", "Full INFO PROMPT"])
    for c in range(1, 5):
        cell = ws2.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for r in rows_sorted:
        page_section = f"{r.page} › {r.section}" if r.section else r.page
        ws2.append([r.code_name, r.feature_name, page_section, r.info_prompt])
    ws2.freeze_panes = "A2"
    autosize(ws2, {1: 12, 2: 38, 3: 38, 4: 90})
    for row_cells in ws2.iter_rows(min_row=2):
        for c in row_cells:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    # ── Sheet 3: Code Name Legend ────────────────────────────────────
    ws3 = wb.create_sheet("Code Name Legend")
    ws3.append(["Prefix", "Page (or area)", "Example", "Notes"])
    for c in range(1, 5):
        cell = ws3.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for page, prefix in PAGE_PREFIXES.items():
        # Find an example code name from rows_sorted.
        example = next((r.code_name for r in rows_sorted if r.page == page), f"{prefix}-1")
        notes = ""
        if prefix == "INF":
            notes = "Reserved for new infrastructure rows that don't appear in SETTINGS_INFO_PROMPTS.txt."
        ws3.append([prefix, page, example, notes])
    autosize(ws3, {1: 10, 2: 36, 3: 14, 4: 70})

    # ── Sheet 4: Glossary / Dropdowns ────────────────────────────────
    ws4 = wb.create_sheet("Glossary")
    ws4.append(["Category", "Value", "Definition"])
    for c in range(1, 4):
        cell = ws4.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    glossary_rows = [
        ("Tier", "T1 (Hard-gate)",
         "Sensitive or destructive: only reachable from System Settings, or from Keyboard Settings when 'Include Advanced Settings in Keyboard Settings' is ON. Restore-from-Backup uses confirm-via-test-mode."),
        ("Tier", "T2 (Visible-but-locked)",
         "Complex setup: visible in Keyboard Settings but requires the Advanced toggle (or a dedicated unlock) before values can be changed. Defaults remain readable."),
        ("Tier", "T3 (Open)",
         "Safe day-to-day controls: always editable in both surfaces."),
        # ── Disclosure (v3, 2026-05-04) ──────────────────────────────
        ("Disclosure", "E (Essential)",
         "Visible to first-time users from day 1; cannot be hidden. The fundamental decisions a user must make to type at all (input method choice, key feedback, basic speech, error beep, the help-icon mechanism)."),
        ("Disclosure", "S (Standard)",
         "Visible by default to everyone. Common day-to-day adjustments — keyboard size, predictions, speak-each-word, basic backup. May be optionally hidden later if a Simplify Mode is added."),
        ("Disclosure", "I (Intermediate)",
         "Hidden until the user opts into 'Show More Options' (future). Power-user adjustments — debounce, auto-repeat, vocabulary promotion, capitalization timing, scan-step delay tuning."),
        ("Disclosure", "A (Advanced)",
         "Hidden until 'Show Advanced Settings' master toggle is ON (List B). Calibration sliders (zone thresholds, response curves), vocabulary frequency-class thresholds, debug overlays, advanced backup, developer settings."),
        # ── Other categories ─────────────────────────────────────────
        ("Surface", "System", "Reachable only from the System Settings activity."),
        ("Surface", "Keyboard", "Reachable only from the Keyboard Settings overlay."),
        ("Surface", "Both", "Reachable from both surfaces (target state for parity)."),
        ("Surface", "Neither", "Not yet exposed in either surface (typically new features)."),
        ("Status", "Implemented per INFO PROMPT",
         "Code matches the INFO PROMPT exactly — only documentation work remains."),
        ("Status", "Behavior gap — needs code update",
         "Implemented but the runtime behavior diverges from the INFO PROMPT (Phase 3D)."),
        ("Status", "Not yet implemented",
         "New setting or feature; needs UI + plumbing in Phase 3 and/or Phase 4."),
        ("Status", "Renamed only", "Behavior is unchanged; the visible label is being updated."),
        ("Audience", "User", "Self-managed device user."),
        ("Audience", "Teacher",
         "Educator setting up JustType for a student; benefits from vocabulary/literacy guidance."),
        ("Audience", "Clinician",
         "Therapist or technician fitting input methods to motor abilities."),
    ]
    for row in glossary_rows:
        ws4.append(row)
    autosize(ws4, {1: 12, 2: 30, 3: 90})
    for row_cells in ws4.iter_rows(min_row=2):
        for c in row_cells:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    # ── Sheet 5: Renames ─────────────────────────────────────────────
    ws5 = wb.create_sheet("Renames")
    ws5.append(["Code Name", "Current Name", "New Name", "Page › Section"])
    for c in range(1, 5):
        cell = ws5.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r in rows_sorted:
        if r.original_name and r.original_name.strip() != r.feature_name.strip():
            ws5.append([
                r.code_name,
                r.original_name,
                r.feature_name,
                f"{r.page} › {r.section}" if r.section else r.page,
            ])
    autosize(ws5, {1: 12, 2: 50, 3: 50, 4: 38})

    # ── Sheet 6: Meta-Instructions ───────────────────────────────────
    ws6 = wb.create_sheet("Meta-Instructions")
    ws6.append(["#", "Directive (verbatim)", "Status"])
    for c in range(1, 4):
        cell = ws6.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    metas = extract_meta_instructions(SOURCE_TXT.read_text(encoding="utf-8", errors="replace"))
    for idx, meta in enumerate(metas, start=1):
        ws6.append([idx, meta, meta_status(meta)])
    autosize(ws6, {1: 6, 2: 110, 3: 42})
    for row_cells in ws6.iter_rows(min_row=2):
        for c in row_cells:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    return wb


META_RE = re.compile(r">>>\s*(.*?)\s*<<<", re.DOTALL)


def extract_meta_instructions(text: str) -> list[str]:
    raw = [clean(m) for m in META_RE.findall(text)]
    # Filter out the noise match from the document's own explanatory note
    # ('Meta-instructions ... are placed between ">>>" and "<<<".' produces
    # a spurious match of `" and "` between the first '>>>' and last '<<<').
    return [m for m in raw if len(m) > 12 and m != '" and "']


# Phase-status mapping for the embedded meta-instructions. Keys are
# substrings of the directive (case-insensitive); the first key found in a
# given directive determines the status. Updated 2026-05-01 to reflect what
# Phases 3A through 3D-Tier-1 have already landed.
META_STATUS_HINTS: list[tuple[str, str]] = [
    # Applied in Phase 3A (renames):
    ("touchscreen switch mode\" should be named",  "Applied (Phase 3A)"),
    # Applied in Phase 3B:
    # (info-icon helper itself — covered by infrastructure row, not a >>> directive)
    # Applied in Phase 3C-A:
    ("disable highlight",                          "Applied (Phase 3C-A)"),
    ("green switch is on the left",                "Applied (Phase 3C-A)"),
    ("highlight timeout",                          "Applied (Phase 3C-A: verified already in correct location)"),
    ("auditory feedback",                          "Applied (pre-Phase 3C: Error Beep already lives on Input Methods page)"),
    # Applied in Phase 3C-B:
    ("alternative input method\" section",         "Applied (Phase 3C-B: verified already in correct order)"),
    ("corner bias",                                "Applied (Phase 3C-B)"),
    ("vocabulary sources",                         "Applied (Phase 3C-A: duplicate Main Vocab Modules table removed)"),
    ("main vocabulary modules",                    "Applied (Phase 3C-A: duplicate table removed)"),
    # Applied in Phase 3D Tier 1:
    ("required code fix",                          "Applied (Phase 3D Tier 1: Δ-35)"),
    # Pending — flagged with hints to which phase will land it:
    ("overlay mode",                               "Pending (Phase 3D Tier 3 research)"),
    ("operation and definition",                   "Pending (Phase 3D Tier 3 research)"),
    ("touchscreen switch height",                  "Applied (Phase 3C-A: UI exists; renderer end-to-end verification still open)"),
    ("info prompts apply to the backup",           "Pending (Phase 3E: Advanced Backup & Restore)"),
    ("new setting]",                               "Mixed — see individual setting status on Features sheet"),
]


def meta_status(directive: str) -> str:
    blob = directive.lower()
    for needle, status in META_STATUS_HINTS:
        if needle in blob:
            return status
    return "Pending"


# ── Main ───────────────────────────────────────────────────────────────────

def main() -> int:
    text = SOURCE_TXT.read_text(encoding="utf-8", errors="replace")
    rows = parse_source(text)

    # Append infrastructure rows.
    for entry in INFRASTRUCTURE_ROWS:
        rows.append(FeatureRow(
            feature_name=entry["feature_name"],
            page=entry["page"],
            section=entry["section"],
            surface="Neither",
            status=entry["status"],
            tier=entry["tier"],
            advanced_affiliation=entry["advanced_affiliation"],
            settings_group=entry["section"],
            tutorial_topic=entry["tutorial_topic"],
            audience=entry["audience"],
            dependencies=entry["dependencies"],
            info_prompt=entry["info_prompt"],
        ))

    apply_known_overrides(rows)
    overlay_phase1_data(rows)
    fill_proposals(rows)
    assign_code_names(rows)

    wb = build_workbook(rows)
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"Wrote {OUTPUT_XLSX}  ({len(rows)} feature rows)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
